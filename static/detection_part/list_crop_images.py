from openpyxl import load_workbook,drawing
import pandas as pd
import os
# dir => directory of images , filepath => existing excel file , classlist => classname
def insert_images(dir,filepath,classlist): #classlist = Total['name']
    images = os.listdir(dir) # read directory
    images = sorted(images, key = lambda x:int(x.split('.')[0])) # sort-filename
    workbook = load_workbook(filepath) # load- excel file
    worksheet = workbook.create_sheet(title='images',index=1) # create worksheet
    i = 1 # cell gap
    j = 0 # index of classname
    for img in images:
        name = img.split('.')[0]
        img = drawing.image.Image(dir+'/'+ img)
        img.width = 100
        img.height = 90
        Anchor = f'A{i}'
        worksheet.add_image(img,anchor=Anchor) # add images
        worksheet[f'C{i+2}'] = name # list name of pics
        worksheet[f'D{i+2}'] = classlist[j] # list classname of pics
        i += 5
        j += 1
    workbook.save(filepath)
# insert_images(dir='../output/crop',filepath='../output/copy.xlsx',classlist=temp)
